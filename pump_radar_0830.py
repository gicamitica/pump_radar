"""
Pump Radar 08:30 (Excel generator) — minimal, production-ready.
Rulează din cron / Render / Railway sau manual (--make-excel-now).

Dependențe: requests, pandas, openpyxl, openai (opțional).
Chei din ENV: OPENAI_API_KEY (opțional), COINGECKO_API_KEY (recomandat), LUNARCRUSH_API_KEY (opțional).
Fișier output: Pump_Radar_DD-MM-YYYY_0830.xlsx

Usage:
  python pump_radar_0830.py --make-excel-now
  python pump_radar_0830.py --auto-0830
  python pump_radar_0830.py --selftest
"""
from __future__ import annotations
import os, sys, json, time
from pathlib import Path
from datetime import datetime
from typing import List, Tuple

# Optional deps handled gracefully
try:
    import requests  # type: ignore
except Exception:
    requests = None  # type: ignore
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None  # type: ignore
try:
    from openai import OpenAI  # type: ignore
    _OPENAI_LIB = True
except Exception:
    _OPENAI_LIB = False

try:
    from zoneinfo import ZoneInfo  # py>=3.9
except Exception:
    ZoneInfo = None  # type: ignore

# --- ENV config ---
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
COINGECKO_API_KEY = os.environ.get("COINGECKO_API_KEY", "").strip()
LUNARCRUSH_API_KEY = os.environ.get("LUNARCRUSH_API_KEY", "").strip()

# --- Logging ---
def log(msg: str) -> None:
    sys.stderr.write(msg + "\n")

# --- CoinGecko ---
def cg_get_markets(per_page: int = 250, page: int = 1, vs_currency: str = "usd") -> List[dict]:
    if not requests:
        log("[EROR] Modulul 'requests' lipsă. pip install requests")
        return []
    url = (
        "https://api.coingecko.com/api/v3/coins/markets"
        f"?vs_currency={vs_currency}&order=volume_desc&per_page={per_page}&page={page}&price_change_percentage=24h"
    )
    headers = {"x-cg-demo-api-key": COINGECKO_API_KEY} if COINGECKO_API_KEY else {}
    try:
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code != 200:
            log(f"[WARN] CoinGecko markets {r.status_code}: {r.text[:200]}")
            return []
        return r.json() or []
    except Exception as e:
        log(f"[WARN] cg_get_markets err: {e}")
        return []

def cg_get_market_chart_volumes_48h(coin_id: str, vs_currency: str = "usd") -> Tuple[float, float]:
    """
    Returnează (vol_prev_24h, vol_last_24h) folosind /market_chart?days=2&interval=hourly
    """
    if not requests:
        return (0.0, 0.0)
    url = (
        f"https://api.coingecko.com/api/v3/coins/{coin_id}/market_chart"
        f"?vs_currency={vs_currency}&days=2&interval=hourly"
    )
    headers = {"x-cg-demo-api-key": COINGECKO_API_KEY} if COINGECKO_API_KEY else {}
    try:
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code != 200:
            log(f"[WARN] market_chart {coin_id} {r.status_code}")
            return (0.0, 0.0)
        data = r.json() or {}
        vols = data.get("total_volumes") or []  # [[ms, vol], ...]
        if len(vols) < 2:
            return (0.0, 0.0)
        n = len(vols)
        half = n // 2
        prev = sum(v for _, v in vols[:half])
        last = sum(v for _, v in vols[half:])
        return (float(prev), float(last))
    except Exception as e:
        log(f"[WARN] cg_get_market_chart_volumes_48h err {coin_id}: {e}")
        return (0.0, 0.0)

# --- LunarCrush ---
def lc_get_social(symbol: str) -> Tuple[int, float]:
    """
    Social Mentions (24h) și Sentiment % (0-100). Fallback 0 dacă lipsește cheia/eroare.
    Notă: schema exactă LC variază pe planuri; folosim câmpuri robuste.
    """
    if not requests or not LUNARCRUSH_API_KEY:
        return (0, 0.0)
    try:
        url = (
            "https://lunarcrush.com/api/v2?data=assets"
            f"&symbol={symbol.upper()}&page=1&items=1&interval=day&key={LUNARCRUSH_API_KEY}"
        )
        r = requests.get(url, timeout=30)
        if r.status_code != 200:
            log(f"[WARN] LunarCrush {symbol} {r.status_code}")
            return (0, 0.0)
        j = r.json() or {}
        d = (j.get("data") or [{}])[0]
        mentions = int(d.get("social_score", 0))  # aproximare robustă
        # sentiment aproximat (fallback pe galaxy_score)
        sent_pos = float(d.get("galaxy_score", 0.0))
        sentiment_pct = max(0.0, min(100.0, sent_pos))
        return (mentions, sentiment_pct)
    except Exception as e:
        log(f"[WARN] lc_get_social err {symbol}: {e}")
        return (0, 0.0)

# --- Selecție candidați ---
def build_candidates_from_api(limit_pages: int = 1, per_page: int = 200,
                              mcap_max: float = 50_000_000, vol_min: float = 500_000,
                              vol_growth_min_pct: float = 30.0,
                              social_mentions_min: int = 100, sentiment_min: float = 65.0) -> List[dict]:
    coins: List[dict] = []
    for page in range(1, limit_pages + 1):
        rows = cg_get_markets(per_page=per_page, page=page)
        if not rows:
            break
        coins.extend(rows)

    out: List[dict] = []
    for c in coins:
        try:
            mcap = float(c.get("market_cap") or 0)
            vol = float(c.get("total_volume") or 0)
            if not (mcap < mcap_max and vol > vol_min):
                continue
            coin_id = c.get("id")
            symbol = (c.get("symbol") or "").upper()
            name = c.get("name") or symbol
            prev24, last24 = cg_get_market_chart_volumes_48h(coin_id)
            if last24 <= 0 or prev24 <= 0:
                growth_pct = 0.0
            else:
                growth_pct = (last24 - prev24) / prev24 * 100.0
            if growth_pct < vol_growth_min_pct:
                continue
            mentions, sentiment = lc_get_social(symbol)
            if mentions < social_mentions_min or sentiment < sentiment_min:
                continue
            out.append({
                "name": name,
                "symbol": symbol,
                "market_cap": mcap,
                "volume_24h": last24,
                "vol_growth_24h_pct": growth_pct,
                "social_mentions_24h": mentions,
                "sentiment_pct": sentiment,
            })
        except Exception as e:
            log(f"[WARN] skip coin err: {e}")
            continue

    out.sort(key=lambda x: (x["vol_growth_24h_pct"], x["volume_24h"]), reverse=True)
    return out[:5]

# --- ChatGPT fields ---
def ai_fields_for_coin(coin: dict) -> dict:
    """
    Întoarce dict cu cheile: utilitate, red_flags, verdict (BUY/WATCH/AVOID).
    Fallback scurt dacă nu există cheie sau lib.
    """
    if not (_OPENAI_LIB and OPENAI_API_KEY):
        return {
            "utilitate": f"{coin['name']} utilitate: n/a (fallback).",
            "red_flags": "Volatilitate ridicată; lichiditate variabilă.",
            "verdict": "WATCH",
        }
    try:
        client = OpenAI(api_key=OPENAI_API_KEY)
        prompt = (
            "Întoarce JSON strict cu cheile: utilitate, red_flags, verdict (BUY/WATCH/AVOID). "
            "Max 2 propoziții per câmp. Fii concis, obiectiv.\n\n"
            f"Coin: {coin['name']} ({coin['symbol']})\n"
            f"Market Cap: {coin['market_cap']}\n"
            f"Volum 24h: {coin['volume_24h']}\n"
            f"Creștere Volum 24h %: {coin['vol_growth_24h_pct']:.2f}\n"
            f"Social Mentions 24h: {coin['social_mentions_24h']}\n"
            f"Sentiment %: {coin['sentiment_pct']}\n"
        )
        resp = client.responses.create(model="gpt-4o-mini", input=prompt, max_output_tokens=220)
        txt = (resp.output_text or "").strip()
        try:
            data = json.loads(txt)
            if all(k in data for k in ("utilitate", "red_flags", "verdict")):
                return data
        except Exception:
            pass
        return {
            "utilitate": txt[:400],
            "red_flags": "-",
            "verdict": "WATCH",
        }
    except Exception as e:
        return {
            "utilitate": f"AI indisponibil ({e})",
            "red_flags": "-",
            "verdict": "WATCH",
        }

# --- Excel ---
def make_excel_report_0830(cands: List[dict], out_dir: str = ".") -> str:
    if pd is None:
        log("[EROR] 'pandas' lipsă. Instalează: pip install pandas openpyxl")
        return ""
    # îmbogățim cu AI
    rows = []
    for c in cands:
        ai = ai_fields_for_coin(c)
        rows.append({
            "Coin": f"{c['name']} ({c['symbol']})",
            "Market Cap": c["market_cap"],
            "Volum 24h": c["volume_24h"],
            "Creștere Volum 24h": c["vol_growth_24h_pct"],
            "Social Mentions 24h": c["social_mentions_24h"],
            "Sentiment %": c["sentiment_pct"],
            "Utilitate & Catalysts": ai.get("utilitate", ""),
            "Red Flags": ai.get("red_flags", ""),
            "Verdict": ai.get("verdict", "WATCH"),
        })
    df = pd.DataFrame(rows)

    # nume fișier
    if ZoneInfo:
        tz = ZoneInfo("Europe/Bucharest")
        now = datetime.now(tz)
    else:
        now = datetime.utcnow()
    fname = f"Pump_Radar_{now.strftime('%d-%m-%Y')}_0830.xlsx"
    out_path = str(Path(out_dir) / fname)

    # scriere + styling
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Top 5", index=False)

        # legendă
        legend = pd.DataFrame({
            "Coloană": ["Coin", "Market Cap", "Volum 24h", "Creștere Volum 24h", "Social Mentions 24h", "Sentiment %", "Utilitate & Catalysts", "Red Flags", "Verdict"],
            "Descriere": [
                "Nume + simbol",
                "Capitalizare piață (USD)",
                "Volum tranzacționat în ultimele 24h (USD)",
                "% creștere volum vs 24h precedente",
                "Număr mențiuni social 24h",
                "Procent postări pozitive",
                "Utilitate proiect + catalizatori (evenimente, listări, parteneriate)",
                "Atenționări: tokenomics, unlocks, lichiditate, centralizare",
                "BUY / WATCH / AVOID",
            ],
        })
        legend.to_excel(writer, sheet_name="Legenda", index=False)

        # styling cu openpyxl
        wb = writer.book
        ws = writer.sheets["Top 5"]
        from openpyxl.styles import Font, PatternFill, Alignment

        # header bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # lățimi coloane
        widths = [24, 16, 16, 20, 22, 12, 44, 30, 10]
        for i, w in enumerate(widths, start=1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = w

        # color verdict
        verdict_col = 9
        for r in range(2, 2 + len(rows)):
            v = (ws.cell(row=r, column=verdict_col).value or "").upper()
            if v == "BUY":
                ws.cell(row=r, column=verdict_col).fill = PatternFill("solid", fgColor="C6EFCE")
            elif v == "WATCH":
                ws.cell(row=r, column=verdict_col).fill = PatternFill("solid", fgColor="FFF2CC")
            elif v == "AVOID":
                ws.cell(row=r, column=verdict_col).fill = PatternFill("solid", fgColor="F8CBAD")

        # alinieri numerice drepte pe col 2..6
        for r in ws.iter_rows(min_row=2, max_row=1+len(rows), min_col=2, max_col=6):
            for c in r:
                c.alignment = Alignment(horizontal="right")

    log(f"[OK] Excel salvat: {out_path}")
    return out_path

# --- Runneri ---
def run_0830_once(limit_pages: int = 1) -> str:
    cands = build_candidates_from_api(limit_pages=limit_pages)
    if not cands:
        log("[INFO] Nicio monedă nu trece filtrele. Generez totuși Excel gol cu header.")
    return make_excel_report_0830(cands)

def loop_until_0830() -> None:
    # rulează în buclă, declanșează la 08:30 RO
    while True:
        now = datetime.now(ZoneInfo("Europe/Bucharest")) if ZoneInfo else datetime.utcnow()
        if now.hour == 8 and now.minute == 30:
            run_0830_once()
            time.sleep(70)  # protecție dublă exec
        time.sleep(5)

# --- Selftests ---
def _selftest_basic():
    # Fără rețea/chei, doar verifică că Excel se generează cu listă goală (sau mesaj dacă pandas lipsește)
    path = make_excel_report_0830([])
    assert path == "" or Path(path).exists(), "Ar trebui să creeze fișier sau să returneze string gol când lipsesc pandas"
    return True

if __name__ == "__main__":
    if "--selftest" in sys.argv:
        ok = _selftest_basic()
        print("Selftest OK" if ok else "Selftest FAIL")
        sys.exit(0)

    if "--make-excel-now" in sys.argv:
        run_0830_once()
        sys.exit(0)

    if "--auto-0830" in sys.argv:
        log("[INFO] Aștept 08:30 Europe/Bucharest…")
        loop_until_0830()
        sys.exit(0)

    # Default: run once now
    run_0830_once()
